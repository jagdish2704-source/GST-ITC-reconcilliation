import os
import re
import sqlite3
import threading
import datetime
import pandas as pd
import traceback
from tkinter import *
from tkinter import ttk, filedialog, messagebox
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from rcm_processing import split_rcm_rows
import shutil
import tempfile
from openpyxl import load_workbook
try:
    from PIL import Image, ImageTk
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False


# ==========================================================
# CONFIG - Permanent Change Heading File (no need to upload each time)
# ==========================================================
# App reads column mapping from this fixed Excel file. Edit this file once to set your mappings.
# Tried in order: 1) Desktop path, 2) change_heading.xlsx in app folder
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
CHANGE_HEADING_PATHS = [
    r"d:\Currnt all data\Desktop\change heading.xlsx",  # Your permanent file - save .pub as .xlsx
    os.path.join(SCRIPT_DIR, "change_heading.xlsx"),    # Fallback: app folder
]

# ==========================================================
# DATABASE (SQLite) - User Management + Audit Trail
# ==========================================================
DB_FILE = "gst_reco_users.db"

def init_db():
    conn = sqlite3.connect(DB_FILE)
    cur = conn.cursor()

    cur.execute("""
    CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE NOT NULL,
        password TEXT NOT NULL,
        role TEXT NOT NULL
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS audit_log (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT,
        timestamp TEXT,
        tolerance TEXT,
        source_file TEXT
    )
    """)

    cur.execute("SELECT COUNT(*) FROM users WHERE username='Admin'")
    if cur.fetchone()[0] == 0:
        cur.execute("INSERT INTO users(username,password,role) VALUES(?,?,?)",
                    ("Admin", "Gima123", "ADMIN"))

    conn.commit()
    conn.close()


def verify_login(username, password):
    conn = sqlite3.connect(DB_FILE)
    cur = conn.cursor()
    cur.execute("SELECT role FROM users WHERE username=? AND password=?", (username, password))
    row = cur.fetchone()
    conn.close()
    return row[0] if row else None


def create_user(username, password, role="USER"):
    conn = sqlite3.connect(DB_FILE)
    cur = conn.cursor()
    try:
        cur.execute("INSERT INTO users(username,password,role) VALUES(?,?,?)",
                    (username, password, role))
        conn.commit()
        return True
    except:
        return False
    finally:
        conn.close()


def delete_user(username):
    conn = sqlite3.connect(DB_FILE)
    cur = conn.cursor()
    cur.execute("DELETE FROM users WHERE username=? AND username<>'Admin'", (username,))
    conn.commit()
    conn.close()


def get_all_users():
    conn = sqlite3.connect(DB_FILE)
    cur = conn.cursor()
    cur.execute("SELECT username, role FROM users")
    rows = cur.fetchall()
    conn.close()
    return rows


def log_audit(username, tolerance, source_file):
    conn = sqlite3.connect(DB_FILE)
    cur = conn.cursor()
    ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cur.execute("INSERT INTO audit_log(username,timestamp,tolerance,source_file) VALUES(?,?,?,?)",
                (username, ts, tolerance, source_file))
    conn.commit()
    conn.close()


# ==========================================================
# NORMALIZATION + VALIDATION
# ==========================================================
GSTIN_REGEX = r"^[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z][1-9A-Z]Z[0-9A-Z]$"

def normalize_invoice(inv):
    if pd.isna(inv):
        return ""
    inv = str(inv).upper().strip()
    inv = inv.replace(" ", "")
    inv = re.sub(r"[\/\-\_\.\s]", "", inv)
    inv = re.sub(r"[^A-Z0-9]", "", inv)
    return inv


def clean_unregistered_value(gstin):
    if pd.isna(gstin):
        return ""
    g = str(gstin).strip().upper()
    g = g.replace("-", "").replace(" ", "")
    return g


def validate_gstin(gstin):
    if pd.isna(gstin):
        return False
    gstin = str(gstin).strip().upper()
    if clean_unregistered_value(gstin) == "UNREGISTERED":
        return True
    return bool(re.match(GSTIN_REGEX, gstin))


# ==========================================================
# EXCEL FORMATTING (openpyxl)
# ==========================================================
def style_worksheet(ws, tab_color="1F4E79"):
    ws.sheet_properties.tabColor = tab_color

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border

    ws.freeze_panes = "A2"

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)


# ==========================================================
# CHANGE IN HEADING - Read mapping from permanent Excel file
# ==========================================================
STANDARD_HEADINGS = frozenset([
    "GSTIN", "Invoice No", "Invoice Date", "Supplier Name",
    "Taxable Value", "Invoice Value", "CGST", "SGST", "IGST"
])


def get_change_heading_path():
    """Return first existing path for the permanent change heading file."""
    for p in CHANGE_HEADING_PATHS:
        if os.path.isfile(p):
            return p
    return None


def create_default_change_heading_file():
    """Create change_heading.xlsx in app folder with template if it doesn't exist."""
    app_path = os.path.join(SCRIPT_DIR, "change_heading.xlsx")
    if os.path.isfile(app_path):
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Change in heading"
    ws.append(["Original Heading (as in Book/2B)", "Standard Heading (target)"])
    ws.append(["Supplier GSTIN", "GSTIN"])
    ws.append(["Invoice Number", "Invoice No"])
    ws.append(["Invoice Date", "Invoice Date"])
    ws.append(["Party Name", "Supplier Name"])
    ws.append(["Taxable Amount", "Taxable Value"])
    ws.append(["Total Value", "Invoice Value"])
    ws.append(["Central Tax", "CGST"])
    ws.append(["State Tax", "SGST"])
    ws.append(["Integrated Tax", "IGST"])
    wb.save(app_path)


def read_change_in_heading():
    """
    Read permanent 'Change in heading' Excel file (no upload needed).
    Supports both formats:
      - Column A = Original (as in Book/2B), Column B = Standard (target)
      - Column A = Standard (target), Column B = Original (as in Book/2B)
    Returns dict: {original_heading: standard_heading}
    """
    path = get_change_heading_path()
    if not path:
        return {}

    try:
        df = pd.read_excel(path, sheet_name=0, header=None, dtype=str)
    except Exception:
        return {}
    if df.empty or len(df.columns) < 2:
        return {}

    # Skip header row if first row contains 'original' or 'standard'
    rename_dict = {}
    start_row = 0
    if len(df) > 0:
        v0 = str(df.iloc[0, 0]).lower() if pd.notna(df.iloc[0, 0]) else ""
        v1 = str(df.iloc[0, 1]).lower() if pd.notna(df.iloc[0, 1]) else ""
        if "original" in v0 or "standard" in v0 or "heading" in v0 or "original" in v1 or "standard" in v1 or "heading" in v1:
            start_row = 1

    for i in range(start_row, len(df)):
        a = str(df.iloc[i, 0]).strip() if pd.notna(df.iloc[i, 0]) else ""
        b = str(df.iloc[i, 1]).strip() if pd.notna(df.iloc[i, 1]) else ""
        if not a or not b:
            continue
        if a in STANDARD_HEADINGS and b not in STANDARD_HEADINGS:
            rename_dict[b] = a
        else:
            rename_dict[a] = b

    return rename_dict


def apply_heading_mapping(df, rename_dict):
    """
    Rename columns using mapping. Uses flexible matching (strip, case-insensitive)
    so 'Supplier GSTIN ' or 'supplier gstin' matches 'Supplier GSTIN' in mapping.
    """
    if not rename_dict:
        return df
    # Build lookup: normalized original -> standard
    norm_to_std = {}
    for orig, std in rename_dict.items():
        norm_to_std[orig.strip().lower()] = std
    # Rename: for each column, match normalized
    cols_to_rename = {}
    for col in df.columns:
        col_str = str(col).strip()
        col_norm = col_str.lower()
        if col_norm in norm_to_std:
            cols_to_rename[col] = norm_to_std[col_norm]
    return df.rename(columns=cols_to_rename)


def strip_df_columns(df):
    """Remove leading/trailing spaces from column names."""
    df = df.copy()
    df.columns = [str(c).strip() if c is not None else "" for c in df.columns]
    return df


def combine_tax_columns(df):
    """
    Combine REC + NREC tax columns if present (e.g. CGST_AMT_REC + CGST_AMT_NREC -> CGST).
    Creates CGST, SGST, IGST if both REC and NREC columns exist.
    """
    df = df.copy()
    def _combine(col_rec, col_nrec, new_name):
        cols_norm = {str(c).strip().lower(): c for c in df.columns}
        rec_norm = col_rec.lower().strip()
        nrec_norm = col_nrec.lower().strip()
        if rec_norm in cols_norm and nrec_norm in cols_norm:
            r = pd.to_numeric(df[cols_norm[rec_norm]], errors="coerce").fillna(0)
            n = pd.to_numeric(df[cols_norm[nrec_norm]], errors="coerce").fillna(0)
            df[new_name] = r + n
        elif rec_norm in cols_norm:
            df[new_name] = pd.to_numeric(df[cols_norm[rec_norm]], errors="coerce").fillna(0)
    _combine("CGST_AMT_REC", "CGST_AMT_NREC", "CGST")
    _combine("SGST_AMT_REC", "SGST_AMT_NREC", "SGST")
    _combine("IGST_AMT_REC", "IGST_AMT_NREC", "IGST")
    return df


def remove_zero_tax_duplicates_from_book(book_df, zero_taxes_df, gstin_col_name, inv_col_name, log=None):
    """
    Remove rows from original Book dataframe that match entries in zero_taxes_df.
    Matching is done on GSTIN (case-insensitive) and normalized Invoice No.
    Returns a filtered copy of book_df. If any error occurs, returns original book_df.
    """
    try:
        if zero_taxes_df is None or zero_taxes_df.empty:
            return book_df.copy()

        # Build set of keys from zero_taxes_df
        def _zero_key(r):
            gst = str(r.get("GSTIN", "")).upper().strip() if pd.notna(r.get("GSTIN", "")) else ""
            inv = normalize_invoice(r.get("Invoice No", ""))
            return gst + "|" + inv

        zero_keys = set(zero_taxes_df.apply(_zero_key, axis=1).tolist())

        def _is_in_zero(row):
            try:
                gst = str(row.get(gstin_col_name, "")).upper().strip() if not pd.isna(row.get(gstin_col_name, "")) else ""
                inv = normalize_invoice(row.get(inv_col_name, ""))
                return (gst + "|" + inv) in zero_keys
            except Exception:
                return False

        filtered = book_df[~book_df.apply(_is_in_zero, axis=1)].copy()
        return filtered
    except Exception as ex:
        if log:
            log(f"remove_zero_tax_duplicates_from_book: failed - {ex}")
        return book_df.copy()


def generate_summary_from_reco(reco_df):
    """
    Build summary dataframe from the Reco Result only.
    Counts distinct invoices per Status (unique GSTIN + normalized Invoice No).
    Returns DataFrame with columns ['Status', 'Count'].
    """
    if reco_df is None or reco_df.empty:
        return pd.DataFrame(columns=["Status", "Count"])

    try:
        tmp = reco_df.copy()
        # Create unique key per invoice
        tmp["__Key"] = tmp.apply(lambda r: (str(r.get("GSTIN", "")).upper().strip() + "|" + normalize_invoice(r.get("Invoice No", ""))), axis=1)
        tmp = tmp.drop_duplicates(subset=["__Key"])
        summary = tmp.groupby("Status").size().reset_index(name="Count")
        # Ensure columns order
        summary = summary[["Status", "Count"]]
        return summary
    except Exception:
        return pd.DataFrame(columns=["Status", "Count"])


def generate_correction_report(reco_df, out_dir, log=None):
    """Create Correction_Report.xlsx with three sheets: Doubt, Partially Matched, Error Report."""
    if reco_df is None:
        raise Exception("Reco DataFrame is None")

    doubt = reco_df[reco_df.get("Remark-2", "").astype(str).str.contains("doubt", case=False, na=False)].copy()
    partial = reco_df[reco_df.get("Status", "").astype(str).str.lower().isin(["partial match", "partially matched"])].copy()
    error = reco_df[(reco_df.get("Status", "").astype(str).str.lower() == "error") |
                    (reco_df.get("Remark-2", "").astype(str).str.contains("error", case=False, na=False))].copy()

    out_path = os.path.join(out_dir, "Correction_Report.xlsx")
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        doubt.to_excel(writer, sheet_name="Doubt", index=False)
        partial.to_excel(writer, sheet_name="Partially Matched", index=False)
        error.to_excel(writer, sheet_name="Error Report", index=False)

    if log:
        log(f"Saved Correction_Report.xlsx to {out_path} (Doubt={len(doubt)}, Partial={len(partial)}, Error={len(error)})")

    return out_path, {"Doubt": doubt, "Partially Matched": partial, "Error Report": error}


def process_uploaded_correction(uploaded_path, original_input_path, reco_df, log=None):
    """Read uploaded correction report, detect changes vs reco_df, apply to a temp copy of original_input_path and highlight changes.

    Returns: (tmp_corrected_path, modifications_list, unmatched_list) or (None, [], unmatched_list) if no modifications.
    """
    try:
        sheets = pd.read_excel(uploaded_path, sheet_name=None, dtype=str)
    except Exception as ex:
        raise Exception(f"Failed reading uploaded file: {ex}")

    # Normalize sheet key map
    sk = {k.lower(): k for k in sheets.keys()}
    for req in ["doubt", "partially matched", "error report"]:
        if req not in sk:
            raise Exception(f"Uploaded file missing required sheet: {req}")

    uploaded = {"Doubt": sheets[sk['doubt']], "Partially Matched": sheets[sk['partially matched']], "Error Report": sheets[sk['error report']]}

    # Build reco lookup
    def mk_key(r):
        gst = str(r.get('GSTIN', '')).upper().strip() if not pd.isna(r.get('GSTIN', '')) else ''
        inv = normalize_invoice(r.get('Invoice No', ''))
        return gst + '|' + inv

    reco_map = {mk_key(row): row for _, row in reco_df.iterrows()}

    modifications = []
    unmatched = []

    for sheet_name, df in uploaded.items():
        for _, urow in df.iterrows():
            # build key from uploaded row (may use corrected invoice)
            key = mk_key(urow)
            orig = reco_map.get(key)
            search_key = key

            # if no direct match, try using original index column if present
            if orig is None and '__Index' in urow:
                try:
                    idx = int(urow.get('__Index', -1))
                    if idx in reco_df.index:
                        orig = reco_df.loc[idx]
                except Exception:
                    pass

            # fallback: try match by GSTIN and invoice date (helps when invoice number changed)
            if orig is None:
                gst_val = str(urow.get('GSTIN', '')).upper().strip()
                date_book = urow.get('Invoice Date_Book', '')
                candidates = reco_df[reco_df.get('GSTIN', '').astype(str).str.upper().str.strip() == gst_val]
                if date_book and not candidates.empty:
                    try:
                        candidates = candidates[candidates.get('Invoice Date_Book', '') == date_book]
                    except Exception:
                        pass
                if not candidates.empty:
                    orig = candidates.iloc[0]

            if orig is None:
                unmatched.append({'sheet': sheet_name, 'key': key})
                continue

            # determine search key from original row (used for locating row in workbook)
            ogst = str(orig.get('GSTIN', '')).upper().strip() if not pd.isna(orig.get('GSTIN', '')) else ''
            oinv = normalize_invoice(orig.get('Invoice No', ''))
            search_key = ogst + '|' + oinv

            for col in df.columns:
                newv = '' if pd.isna(urow.get(col, '')) else str(urow.get(col, '')).strip()
                oldv = '' if pd.isna(orig.get(col, '')) else str(orig.get(col, '')).strip()
                if newv != oldv:
                    target = 'Both'
                    col_base = col
                    if col.endswith('_Book'):
                        target = 'Book'
                        col_base = col[:-5]
                    elif col.endswith('_2B') or col.endswith('_2b'):
                        target = 'GSTR2B'
                        col_base = col[:-3]

                    modifications.append({
                        'key': key,
                        'search_key': search_key,
                        'sheet': sheet_name,
                        'column': col,
                        'column_base': col_base,
                        'old': oldv,
                        'new': newv,
                        'target': target
                    })

    if not modifications:
        return None, [], unmatched

    # Copy original input to temp and apply modifications with highlighting
    base_dir = os.path.dirname(original_input_path)
    fd, tmp_path = tempfile.mkstemp(suffix='_corrected.xlsx', dir=base_dir)
    os.close(fd)
    shutil.copy2(original_input_path, tmp_path)

    wb = load_workbook(tmp_path)

    # find book and gstr sheet names
    book_sheet = None
    gstr_sheet = None
    for s in wb.sheetnames:
        sl = s.lower()
        if 'book' in sl and book_sheet is None:
            book_sheet = s
        if ('2b' in sl or 'gstr2b' in sl) and gstr_sheet is None:
            gstr_sheet = s

    yellow = PatternFill(start_color='FFFF00', fill_type='solid')

    def header_map(ws):
        return {str(cell.value).strip(): cell.column_letter for cell in ws[1] if cell.value is not None}

    if book_sheet:
        bws = wb[book_sheet]
        bh = header_map(bws)
    else:
        bws = None
        bh = {}

    if gstr_sheet:
        gws = wb[gstr_sheet]
        gh = header_map(gws)
    else:
        gws = None
        gh = {}

    # Apply mods
    for m in modifications:
        # use search_key (original position) to locate the row, but keep key for logging if needed
        gst, inv = m.get('search_key', m['key']).split('|')
        col_base = m['column_base']
        newv = m['new']
        targets = [m['target']] if m['target'] != 'Both' else ['Book', 'GSTR2B']

        for t in targets:
            ws = bws if t == 'Book' else gws
            hdr = bh if t == 'Book' else gh
            if ws is None:
                continue

            # locate gst and invoice columns
            gst_col = None
            inv_col = None
            for h, letter in hdr.items():
                ln = h.strip().lower()
                if 'gstin' in ln:
                    gst_col = letter
                if 'invoice' in ln:
                    inv_col = letter

            if not gst_col or not inv_col:
                unmatched.append({'mod': m, 'reason': 'gst/inv col not found in target sheet'})
                continue

            for row in range(2, ws.max_row + 1):
                cell_gst = str(ws[f"{gst_col}{row}"].value or '').upper().strip()
                cell_inv = normalize_invoice(ws[f"{inv_col}{row}"].value or '')
                if cell_gst == gst and cell_inv == inv:
                    # find target column
                    tgt = None
                    for h, letter in hdr.items():
                        if h.strip().lower() == col_base.strip().lower():
                            tgt = letter
                            break
                    if not tgt:
                        for h, letter in hdr.items():
                            if col_base.strip().lower() in h.strip().lower():
                                tgt = letter
                                break
                    if not tgt:
                        unmatched.append({'mod': m, 'reason': 'target col not found'})
                        break

                    ws[f"{tgt}{row}"].value = newv
                    ws[f"{tgt}{row}"].fill = yellow
                    break

    wb.save(tmp_path)
    return tmp_path, modifications, unmatched


# ==========================================================
# CORE RECONCILIATION ENGINE
# ==========================================================
def reconcile(file_path, taxpayer_gstin, period, tolerance, username, progress_callback=None, log_callback=None):
    errors = []

    def log(msg):
        if log_callback:
            log_callback(msg)

    def update_progress(val):
        if progress_callback:
            progress_callback(val)

    update_progress(5)
    log("Reading Excel file...")

    xls = pd.ExcelFile(file_path)
    sheet_names = xls.sheet_names

    book_sheet = None
    gstr2b_sheet = None

    for s in sheet_names:
        s_low = s.lower()
        if "book" in s_low:
            book_sheet = s
        if "2b" in s_low or "gstr2b" in s_low:
            gstr2b_sheet = s

    if not book_sheet or not gstr2b_sheet:
        raise Exception("Excel must contain sheet name containing 'Book' and one containing '2B' / 'GSTR2B'.")

    book_raw = pd.read_excel(file_path, sheet_name=book_sheet, dtype=str)
    gstr_raw = pd.read_excel(file_path, sheet_name=gstr2b_sheet, dtype=str)
    # Strip leading/trailing spaces from column names
    book_raw = strip_df_columns(book_raw)
    gstr_raw = strip_df_columns(gstr_raw)

    # --- STEP 1: Apply "Change in heading" mapping (permanent file - no upload needed) ---
    update_progress(10)
    log("Applying column mapping from permanent 'Change in heading' file...")
    heading_map = read_change_in_heading()
    if heading_map:
        book_raw = apply_heading_mapping(book_raw, heading_map)
        gstr_raw = apply_heading_mapping(gstr_raw, heading_map)
        log(f"Applied {len(heading_map)} column renames from permanent Change in heading.")
    else:
        log("No permanent Change in heading file found; using default column mapping.")

    # Combine REC+NREC tax columns (CGST_AMT_REC + CGST_AMT_NREC -> CGST) if present
    book_raw = combine_tax_columns(book_raw)
    gstr_raw = combine_tax_columns(gstr_raw)

    # --- RCM split (module) ---
    # Identify rows with CGST_AMT_NREC > 0 OR IGST_AMT_NREC > 0, move them to RCM,
    # mark Status='RCM', and exclude from reconciliation.
    book_raw, rcm_raw = split_rcm_rows(book_raw, log=log)

    book_df = book_raw.copy()
    gstr2b_df = gstr_raw.copy()

    update_progress(15)
    log("Cleaning & standardizing columns...")

    required_cols = {
        "GSTIN": ["GSTIN", "Supplier GSTIN", "GSTIN/UIN", "GSTIN_NO", "GSTIN of supplier", "GSTIN of Supplier"],
        "Invoice No": ["Invoice No", "Invoice Number", "Inv No", "Invoice", "INVOICE_NO", "Invoice number"],
        "Invoice Date": ["Invoice Date", "Inv Date", "Date", "INVOICE_DATE"],
        "Supplier Name": ["Supplier Name", "Vendor Name", "Party Name", "SUPP_NAME", "Trade/Legal name"],
        "Taxable Value": ["Taxable Value", "Taxable", "Taxable Amount", "TAXABLE_AMT", "Taxable Value (₹)", "Taxable Value (Rs)"],
        "Invoice Value": ["Invoice Value", "Total Invoice Value", "Total", "INVOICE_AMT", "Invoice Value(₹)", "Invoice Value (Rs)"],
        "CGST": ["CGST", "Central Tax", "CGST_AMT_REC", "CGST_AMT_NREC", "Central Tax(₹)", "Central Tax (Rs)"],
        "SGST": ["SGST", "State Tax", "SGST_AMT_REC", "SGST_AMT_NREC", "State/UT Tax(₹)", "State Tax (Rs)"],
        "IGST": ["IGST", "Integrated Tax", "IGST_AMT_REC", "IGST_AMT_NREC", "Integrated Tax(₹)", "Integrated Tax (Rs)"]
    }

    def map_columns(df):
        # Use flexible matching (strip, case-insensitive) for column names
        df_cols_norm = {str(c).strip().lower(): c for c in df.columns if c is not None}
        mapped = {}
        for std, variants in required_cols.items():
            for v in variants:
                v_norm = str(v).strip().lower()
                if v_norm in df_cols_norm:
                    mapped[std] = df_cols_norm[v_norm]
                    break
        return mapped

    book_map = map_columns(book_df)
    gstr_map = map_columns(gstr2b_df)

    missing_book = [c for c in required_cols.keys() if c not in book_map]
    missing_gstr = [c for c in required_cols.keys() if c not in gstr_map]

    if missing_book:
        raise Exception(f"Missing columns in Book sheet: {missing_book}")
    if missing_gstr:
        raise Exception(f"Missing columns in GSTR2B sheet: {missing_gstr}")

    book_df = book_df.rename(columns={book_map[k]: k for k in book_map})
    gstr2b_df = gstr2b_df.rename(columns={gstr_map[k]: k for k in gstr_map})

    update_progress(25)
    log("Converting numeric columns...")

    num_cols = ["Taxable Value", "Invoice Value", "CGST", "SGST", "IGST"]

    for c in num_cols:
        book_df[c] = pd.to_numeric(book_df[c], errors="coerce").fillna(0)
        gstr2b_df[c] = pd.to_numeric(gstr2b_df[c], errors="coerce").fillna(0)

    book_df["GSTIN"] = book_df["GSTIN"].fillna("").astype(str).str.upper().str.strip()
    gstr2b_df["GSTIN"] = gstr2b_df["GSTIN"].fillna("").astype(str).str.upper().str.strip()

    log("Validating GSTIN format...")

    for idx, gstin in enumerate(book_df["GSTIN"]):
        gst_clean = clean_unregistered_value(gstin)

        if gst_clean == "UNREGISTERED":
            continue

        if gstin.strip() != "" and not validate_gstin(gstin):
            errors.append({
                "Sheet Name": book_sheet,
                "Row Number": idx + 2,
                "GSTIN Value": gstin,
                "Error Type": "Invalid GSTIN Format",
                "Error Description": "GSTIN does not match regex pattern"
            })

    update_progress(35)
    log("Normalizing invoice numbers...")

    book_df["Inv_Normalized"] = book_df["Invoice No"].apply(normalize_invoice)
    gstr2b_df["Inv_Normalized"] = gstr2b_df["Invoice No"].apply(normalize_invoice)

    book_df["Last3"] = book_df["Inv_Normalized"].apply(lambda x: x[-3:] if len(x) >= 3 else x)
    gstr2b_df["Last3"] = gstr2b_df["Inv_Normalized"].apply(lambda x: x[-3:] if len(x) >= 3 else x)

    log("Handling duplicates in Book...")
    book_group_cols = ["GSTIN", "Inv_Normalized"]

    def agg_invoice_value(group):
        inv_vals = group["Invoice Value"].unique()
        if len(inv_vals) == 1:
            return inv_vals[0]
        else:
            return group["Invoice Value"].sum()

    book_agg = book_df.groupby(book_group_cols).agg({
        "Supplier Name": "first",
        "Invoice No": "first",
        "Invoice Date": "first",
        "Taxable Value": "sum",
        "CGST": "sum",
        "SGST": "sum",
        "IGST": "sum"
    }).reset_index()

    invoice_value_series = book_df.groupby(book_group_cols).apply(agg_invoice_value).reset_index(name="Invoice Value")
    book_agg = book_agg.merge(invoice_value_series, on=book_group_cols, how="left")

    book_agg["Last3"] = book_agg["Inv_Normalized"].apply(lambda x: x[-3:] if len(x) >= 3 else x)
    book_df = book_agg.copy()

    update_progress(45)
    log("Preparing matching keys...")

    book_df["MatchKey"] = book_df["GSTIN"] + "|" + book_df["Inv_Normalized"]
    gstr2b_df["MatchKey"] = gstr2b_df["GSTIN"] + "|" + gstr2b_df["Inv_Normalized"]

    log("Merging Book with GSTR2B...")
    merged = book_df.merge(gstr2b_df, on="MatchKey", how="outer", suffixes=("_Book", "_2B"), indicator=True)

    update_progress(60)
    log("Computing differences and status...")

    for c in num_cols:
        merged[f"{c}_Book"] = merged.get(f"{c}_Book", 0).fillna(0)
        merged[f"{c}_2B"] = merged.get(f"{c}_2B", 0).fillna(0)

    merged["GSTIN"] = merged["GSTIN_Book"].where(merged["GSTIN_Book"].notna(), merged["GSTIN_2B"])
    merged["Invoice No"] = merged["Invoice No_Book"].where(merged["Invoice No_Book"].notna(), merged["Invoice No_2B"])

    merged["Supplier Name"] = merged["Supplier Name_Book"].where(
        merged["Supplier Name_Book"].notna() & (merged["Supplier Name_Book"] != ""),
        merged["Supplier Name_2B"]
    )

    merged["Taxable Diff"] = merged["Taxable Value_Book"] - merged["Taxable Value_2B"]
    merged["CGST Diff"] = merged["CGST_Book"] - merged["CGST_2B"]
    merged["SGST Diff"] = merged["SGST_Book"] - merged["SGST_2B"]
    merged["IGST Diff"] = merged["IGST_Book"] - merged["IGST_2B"]

    tol = float(tolerance)

    def determine_status(row):
        if row["_merge"] == "left_only":
            return "Invoice Not Found in 2B"
        if row["_merge"] == "right_only":
            return "Invoice Not Found in Book"

        if (abs(row["Taxable Diff"]) <= tol and
            abs(row["CGST Diff"]) <= tol and
            abs(row["SGST Diff"]) <= tol and
            abs(row["IGST Diff"]) <= tol):
            return "Matched"

        return "Partial Match"

    merged["Status"] = merged.apply(determine_status, axis=1)

    # ==========================================================
    # FIXED DATE REMARK LOGIC (REAL DATE COMPARISON)
    # ==========================================================
    def determine_remark(row):
        remark = ""

        if row["_merge"] == "both":
            d1 = row.get("Invoice Date_Book", "")
            d2 = row.get("Invoice Date_2B", "")

            dt1 = pd.to_datetime(d1, errors="coerce", dayfirst=True)
            dt2 = pd.to_datetime(d2, errors="coerce", dayfirst=True)

            if pd.notna(dt1) and pd.notna(d2):
                if dt1.date() != dt2.date():
                    remark += "Invoice Date mismatch. "

        return remark.strip()

    merged["Remark"] = merged.apply(determine_remark, axis=1)

    # ==========================================================
    # DOUBT LOGIC (PAIR BASED)
    # ==========================================================
    update_progress(70)
    log("Applying DOUBT logic...")

    merged["Remark-2"] = ""

    merged["InvNorm_Book"] = merged["Invoice No_Book"].apply(normalize_invoice)
    merged["InvNorm_2B"] = merged["Invoice No_2B"].apply(normalize_invoice)

    merged["Last3_Book"] = merged["InvNorm_Book"].apply(lambda x: x[-3:] if len(x) >= 3 else x)
    merged["Last3_2B"] = merged["InvNorm_2B"].apply(lambda x: x[-3:] if len(x) >= 3 else x)

    left_only = merged[merged["_merge"] == "left_only"].copy()
    right_only = merged[merged["_merge"] == "right_only"].copy()

    left_only["Key"] = left_only["GSTIN"].astype(str) + "|" + left_only["Last3_Book"].astype(str)
    right_only["Key"] = right_only["GSTIN"].astype(str) + "|" + right_only["Last3_2B"].astype(str)

    common_keys = set(left_only["Key"]).intersection(set(right_only["Key"]))

    def doubt_marker(row):
        if row["_merge"] == "left_only":
            k = str(row["GSTIN"]) + "|" + str(row["Last3_Book"])
            if k in common_keys:
                return "Doubt"

        if row["_merge"] == "right_only":
            k = str(row["GSTIN"]) + "|" + str(row["Last3_2B"])
            if k in common_keys:
                return "Doubt"

        return ""

    merged["Remark-2"] = merged.apply(doubt_marker, axis=1)

    update_progress(75)
    log("Preparing Reco Result + Zero Taxes separation...")

    output_cols = [
        "GSTIN", "Supplier Name", "Invoice No",
        "Invoice Date_Book", "Invoice Date_2B",
        "Taxable Value_Book", "Taxable Value_2B", "Taxable Diff",
        "CGST_Book", "CGST_2B", "CGST Diff",
        "SGST_Book", "SGST_2B", "SGST Diff",
        "IGST_Book", "IGST_2B", "IGST Diff",
        "Invoice Value_Book", "Invoice Value_2B",
        "Status", "Remark", "Remark-2"
    ]

    for c in output_cols:
        if c not in merged.columns:
            merged[c] = ""

    reco_result_full = merged[output_cols].copy()

    zero_taxes = reco_result_full[
        (reco_result_full["Status"] == "Invoice Not Found in 2B") &
        (reco_result_full["CGST_Book"] == 0) &
        (reco_result_full["SGST_Book"] == 0) &
        (reco_result_full["IGST_Book"] == 0)
    ].copy()

    reco_result = reco_result_full.drop(zero_taxes.index).copy()

    update_progress(80)
    log("Preparing Summary & Error Report...")

    # Summary should count statuses from the Reco Result (excluding Zero Taxes)
    # Use distinct invoice count per Status from Reco Result only
    summary = generate_summary_from_reco(reco_result)

    zero_taxes_count = len(zero_taxes)

    summary_meta = pd.DataFrame([
        ["GSTIN of Taxpayer", taxpayer_gstin],
        ["Reconciliation Period", period],
        ["Username", username],
        ["Timestamp", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["Tolerance Used", tolerance],
        ["Source File", os.path.basename(file_path)],
        ["Zero Taxes Count", zero_taxes_count]
    ], columns=["Field", "Value"])

    error_report = pd.DataFrame(errors)

    # ==========================================================
    # UPDATE ORIGINAL BOOK & 2B SHEETS WITH STATUS COLUMN
    # ==========================================================
    update_progress(85)
    log("Updating original Book and GSTR2B sheets with Status column...")

    status_map_book = merged[merged["_merge"] != "right_only"].copy()
    status_map_book["GSTIN_KEY"] = status_map_book["GSTIN_Book"].fillna("").astype(str).str.upper().str.strip()
    status_map_book["INV_KEY"] = status_map_book["Invoice No_Book"].apply(normalize_invoice)
    status_map_book["KEY"] = status_map_book["GSTIN_KEY"] + "|" + status_map_book["INV_KEY"]
    status_dict_book = dict(zip(status_map_book["KEY"], status_map_book["Status"]))

    status_map_2b = merged[merged["_merge"] != "left_only"].copy()
    status_map_2b["GSTIN_KEY"] = status_map_2b["GSTIN_2B"].fillna("").astype(str).str.upper().str.strip()
    status_map_2b["INV_KEY"] = status_map_2b["Invoice No_2B"].apply(normalize_invoice)
    status_map_2b["KEY"] = status_map_2b["GSTIN_KEY"] + "|" + status_map_2b["INV_KEY"]
    status_dict_2b = dict(zip(status_map_2b["KEY"], status_map_2b["Status"]))

    def find_column(df, candidates):
        for c in candidates:
            if c in df.columns:
                return c
        return None

    book_gstin_col = find_column(book_raw, required_cols["GSTIN"])
    book_inv_col = find_column(book_raw, required_cols["Invoice No"])

    gstr_gstin_col = find_column(gstr_raw, required_cols["GSTIN"])
    gstr_inv_col = find_column(gstr_raw, required_cols["Invoice No"])

    if not book_gstin_col or not book_inv_col:
        raise Exception("Original Book sheet GSTIN/Invoice columns not found.")
    if not gstr_gstin_col or not gstr_inv_col:
        raise Exception("Original GSTR2B sheet GSTIN/Invoice columns not found.")

    if "Status" not in book_raw.columns:
        book_raw["Status"] = ""
    if "Status" not in gstr_raw.columns:
        gstr_raw["Status"] = ""

    def get_book_status(row):
        gstin = str(row[book_gstin_col]).upper().strip() if not pd.isna(row[book_gstin_col]) else ""
        inv = normalize_invoice(row[book_inv_col])
        key = gstin + "|" + inv
        return status_dict_book.get(key, "")

    book_raw["Status"] = book_raw.apply(get_book_status, axis=1)

    def get_2b_status(row):
        gstin = str(row[gstr_gstin_col]).upper().strip() if not pd.isna(row[gstr_gstin_col]) else ""
        inv = normalize_invoice(row[gstr_inv_col])
        key = gstin + "|" + inv
        return status_dict_2b.get(key, "")

    gstr_raw["Status"] = gstr_raw.apply(get_2b_status, axis=1)

    # ==========================================================
    # WRITE OUTPUT EXCEL
    # ==========================================================
    update_progress(90)
    log("Writing Excel output with formatting...")

    wb = Workbook()
    wb.remove(wb.active)

    # -------------------- Reco Result Sheet --------------------
    ws1 = wb.create_sheet("Reco Result")

    meta_lines = [
        ["GSTIN of Taxpayer", taxpayer_gstin],
        ["Reconciliation Period", period],
        ["Username", username],
        ["Timestamp", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["Tolerance Used", tolerance],
        ["Source File", os.path.basename(file_path)]
    ]

    for line in meta_lines:
        ws1.append(line)

    ws1.append([""])

    for r in dataframe_to_rows(reco_result, index=False, header=True):
        ws1.append(r)

    for r in range(1, len(meta_lines) + 1):
        ws1[f"A{r}"].font = Font(bold=True, color="1F4E79")

    header_row = len(meta_lines) + 2

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws1[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws1.freeze_panes = f"A{header_row + 1}"

    red_fill = PatternFill("solid", fgColor="FFC7CE")
    status_col_index = list(reco_result.columns).index("Status") + 1

    for r in range(header_row + 1, ws1.max_row + 1):
        status_val = ws1.cell(row=r, column=status_col_index).value
        if status_val in ["Partial Match", "Invoice Not Found in 2B", "Invoice Not Found in Book"]:
            for c in range(1, ws1.max_column + 1):
                ws1.cell(row=r, column=c).fill = red_fill

    for col in ws1.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws1.column_dimensions[col_letter].width = min(max_len + 2, 50)

    ws1.sheet_properties.tabColor = "1F4E79"

    # -------------------- Zero Taxes Sheet --------------------
    ws2 = wb.create_sheet("Zero Taxes")
    for r in dataframe_to_rows(zero_taxes, index=False, header=True):
        ws2.append(r)
    style_worksheet(ws2, tab_color="00B0F0")

    # -------------------- Summary Sheet --------------------
    ws3 = wb.create_sheet("Summary")
    ws3.append(["Summary Information"])
    ws3.append([""])
    for r in dataframe_to_rows(summary_meta, index=False, header=True):
        ws3.append(r)
    ws3.append([""])
    for r in dataframe_to_rows(summary, index=False, header=True):
        ws3.append(r)
    style_worksheet(ws3, tab_color="92D050")

    # -------------------- Error Report Sheet --------------------
    ws4 = wb.create_sheet("Error Report")
    if not error_report.empty:
        for r in dataframe_to_rows(error_report, index=False, header=True):
            ws4.append(r)
    else:
        ws4.append(["No errors found"])
    style_worksheet(ws4, tab_color="FF0000")

    # -------------------- Original Book Sheet with Status --------------------
    # Remove Book rows present in Zero Taxes (based on GSTIN + Invoice No)
    book_out_for_ws = remove_zero_tax_duplicates_from_book(book_raw, zero_taxes, book_gstin_col, book_inv_col, log=log)

    ws5 = wb.create_sheet(book_sheet)
    for r in dataframe_to_rows(book_out_for_ws, index=False, header=True):
        ws5.append(r)
    style_worksheet(ws5, tab_color="FFD966")

    # -------------------- RCM Sheet (moved rows) --------------------
    ws7 = wb.create_sheet("RCM")
    rcm_out = rcm_raw if "rcm_raw" in locals() and rcm_raw is not None else None
    if rcm_out is None:
        rcm_out = book_raw.iloc[0:0].copy()
    if rcm_out.empty:
        # Write headers only (same as Book) when no RCM rows
        for r in dataframe_to_rows(book_raw.iloc[0:0].copy(), index=False, header=True):
            ws7.append(r)
    else:
        for r in dataframe_to_rows(rcm_out, index=False, header=True):
            ws7.append(r)
    style_worksheet(ws7, tab_color="FF9900")

    # -------------------- Original GSTR2B Sheet with Status --------------------
    ws6 = wb.create_sheet(gstr2b_sheet)
    for r in dataframe_to_rows(gstr_raw, index=False, header=True):
        ws6.append(r)
    style_worksheet(ws6, tab_color="C6E0B4")

    output_file = os.path.join(
        os.path.dirname(file_path),
        f"GST_ITC_Reco_Output_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )

    wb.save(output_file)

    update_progress(100)
    log("Reconciliation completed successfully.")

    log_audit(username, tolerance, os.path.basename(file_path))

    return output_file, reco_result


# ==========================================================
# GUI APP (Tkinter Enterprise UI)
# ==========================================================
class GSTRecoApp:
    def __init__(self, root):
        self.root = root
        self.root.title("GST ITC Reconciliation - Enterprise Edition")
        self.root.geometry("1400x800")
        self.root.configure(bg="#FFFFFF")

        self.current_user = None
        self.current_role = None
        self.selected_file = None
        self.result_df = None
        self.output_file_path = None
        self.draft_output_path = None
        # correction workflow removed

        self.build_login_ui()

    def build_login_ui(self):
        for w in self.root.winfo_children():
            w.destroy()

        # Main container with gradient-like effect
        main_frame = Frame(self.root, bg="#FFFFFF")
        main_frame.pack(fill=BOTH, expand=True)

        # Header bar with branding
        header_frame = Frame(main_frame, bg="#1F4E79", height=100)
        header_frame.pack(fill=X, side=TOP)
        header_frame.pack_propagate(False)

        # Logo placeholder and title in header
        logo_frame = Frame(header_frame, bg="#1F4E79")
        logo_frame.pack(side=LEFT, padx=20, pady=15)
        
        # Try to load logo image with proper PIL fallback
        logo_loaded = False
        try:
            if PIL_AVAILABLE:
                logo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "gst_logo.png")
                if os.path.isfile(logo_path) and os.path.getsize(logo_path) > 100:
                    img = Image.open(logo_path)
                    img.thumbnail((70, 70), Image.Resampling.LANCZOS)
                    photo = ImageTk.PhotoImage(img)
                    logo_label = Label(logo_frame, image=photo, bg="#1F4E79")
                    logo_label.image = photo
                    logo_label.pack(side=LEFT, padx=5)
                    logo_loaded = True
        except Exception as e:
            pass
        
        # Fallback if logo not loaded
        if not logo_loaded:
            logo_label = Label(logo_frame, text="G", font=("Segoe UI", 48, "bold"), bg="#1F4E79", fg="#0099CC")
            logo_label.pack(side=LEFT, padx=5)

        # Title and subtitle in header
        title_frame = Frame(header_frame, bg="#1F4E79")
        title_frame.pack(side=LEFT, padx=20, pady=15)
        
        Label(title_frame, text="GST ITC Reconciliation", font=("Segoe UI", 24, "bold"),
              bg="#1F4E79", fg="white").pack(anchor=W)
        Label(title_frame, text="Prepared by CA Jitendra Varma", font=("Segoe UI", 10, "bold"),
              bg="#1F4E79", fg="#FFD700").pack(anchor=W)

        # Central login form
        center_frame = Frame(main_frame, bg="#FFFFFF")
        center_frame.pack(fill=BOTH, expand=True, padx=40, pady=60)

        # Login card with subtle shadow effect
        card_frame = Frame(center_frame, bg="white", relief=RAISED, bd=1)
        card_frame.pack(pady=20, padx=20, fill=BOTH, expand=False, side=TOP)

        # Card title
        Label(card_frame, text="Enterprise Login", font=("Segoe UI", 16, "bold"),
              bg="white", fg="#1F4E79").pack(pady=15)

        # User ID field
        Label(card_frame, text="User ID", font=("Segoe UI", 11, "bold"),
              bg="white", fg="#333333").pack(anchor=W, padx=30, pady=(10, 2))
        self.username_entry = Entry(card_frame, font=("Segoe UI", 11), width=35, relief=SOLID, bd=1)
        self.username_entry.pack(padx=30, pady=(2, 15), ipady=8)

        # Password field
        Label(card_frame, text="Password", font=("Segoe UI", 11, "bold"),
              bg="white", fg="#333333").pack(anchor=W, padx=30, pady=(10, 2))
        self.password_entry = Entry(card_frame, show="•", font=("Segoe UI", 11), width=35, relief=SOLID, bd=1)
        self.password_entry.pack(padx=30, pady=(2, 25), ipady=8)

        # Login button with hover effect
        btn_frame = Frame(card_frame, bg="white")
        btn_frame.pack(pady=20)
        
        login_btn = Button(btn_frame, text="Login", font=("Segoe UI", 12, "bold"),
                          bg="#1F4E79", fg="white", width=20, height=2,
                          command=self.login, cursor="hand2", relief=FLAT)
        login_btn.pack()

        Label(card_frame, text="", bg="white").pack(pady=10)

        # Footer
        footer_frame = Frame(main_frame, bg="#F5F5F5", height=40)
        footer_frame.pack(fill=X, side=BOTTOM)
        footer_frame.pack_propagate(False)
        
        Label(footer_frame, text="© 2026 GST ITC Reconciliation System - All Rights Reserved",
              font=("Segoe UI", 9), bg="#F5F5F5", fg="#666666").pack(pady=10)

    def login(self):
        user = self.username_entry.get().strip()
        pwd = self.password_entry.get().strip()

        role = verify_login(user, pwd)
        if role:
            self.current_user = user
            self.current_role = role
            self.build_dashboard_ui()
        else:
            messagebox.showerror("Login Failed", "Invalid User ID or Password.\nTry again.")

    def build_dashboard_ui(self):
        for w in self.root.winfo_children():
            w.destroy()

        # Professional header with branding
        header_frame = Frame(self.root, bg="#1F4E79", height=90)
        header_frame.pack(fill=X, side=TOP)
        header_frame.pack_propagate(False)

        # Left side: Logo and title
        left_header = Frame(header_frame, bg="#1F4E79")
        left_header.pack(side=LEFT, padx=15, pady=10)
        
        # Try to load logo image with proper PIL handling
        logo_loaded = False
        try:
            if PIL_AVAILABLE:
                logo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "gst_logo.png")
                if os.path.isfile(logo_path) and os.path.getsize(logo_path) > 100:
                    img = Image.open(logo_path)
                    img.thumbnail((60, 60), Image.Resampling.LANCZOS)
                    photo = ImageTk.PhotoImage(img)
                    logo_label = Label(left_header, image=photo, bg="#1F4E79")
                    logo_label.image = photo
                    logo_label.pack(side=LEFT, padx=5)
                    logo_loaded = True
        except Exception as e:
            pass
        
        # Fallback if logo not loaded
        if not logo_loaded:
            logo_label = Label(left_header, text="G", font=("Segoe UI", 40, "bold"), bg="#1F4E79", fg="#0099CC")
            logo_label.pack(side=LEFT, padx=5)
        
        title_frame = Frame(left_header, bg="#1F4E79")
        title_frame.pack(side=LEFT, padx=10)
        
        Label(title_frame, text="GST ITC Reconciliation", font=("Segoe UI", 16, "bold"),
              bg="#1F4E79", fg="white").pack(anchor=W)
        Label(title_frame, text="Prepared by CA Jitendra Varma", font=("Segoe UI", 9, "bold"),
              bg="#1F4E79", fg="#FFD700").pack(anchor=W)

        # Right side: User info and Logout
        right_header = Frame(header_frame, bg="#1F4E79")
        right_header.pack(side=RIGHT, padx=15, pady=10)
        
        Label(right_header, text=f"Logged in as: {self.current_user} ({self.current_role})",
              bg="#1F4E79", fg="white", font=("Segoe UI", 10, "bold")).pack(side=LEFT, padx=10)
        Button(right_header, text="Logout", bg="#FF6B6B", fg="white",
               font=("Segoe UI", 10, "bold"), command=self.build_login_ui, relief=FLAT,
               cursor="hand2").pack(side=LEFT, padx=5, pady=5)

        # Main content area
        main_frame = Frame(self.root, bg="#F8F9FA")
        main_frame.pack(fill=BOTH, expand=True, padx=15, pady=15)

        input_frame = LabelFrame(main_frame, text="Input Section", font=("Segoe UI", 12, "bold"),
                                 bg="white", fg="#1F4E79", padx=15, pady=12, relief=SOLID, bd=1)
        input_frame.pack(fill=X, pady=(0, 15))

        Label(input_frame, text="Taxpayer GSTIN:", bg="white", font=("Segoe UI", 10, "bold")).grid(row=0, column=0, sticky=W, pady=8)
        self.taxpayer_entry = Entry(input_frame, font=("Segoe UI", 10), width=30, relief=SOLID, bd=1)
        self.taxpayer_entry.grid(row=0, column=1, padx=10, pady=8, ipady=5)

        Label(input_frame, text="Reconciliation Period:", bg="white", font=("Segoe UI", 10, "bold")).grid(row=0, column=2, sticky=W, padx=(20, 0), pady=8)
        self.period_entry = Entry(input_frame, font=("Segoe UI", 10), width=20, relief=SOLID, bd=1)
        self.period_entry.grid(row=0, column=3, padx=10, pady=8, ipady=5)

        Label(input_frame, text="Tolerance:", bg="white", font=("Segoe UI", 10, "bold")).grid(row=1, column=0, sticky=W, pady=8)
        self.tolerance_var = StringVar(value="2")
        self.tolerance_combo = ttk.Combobox(input_frame, textvariable=self.tolerance_var,
                                            values=["2", "5", "10"], width=10, state="readonly")
        self.tolerance_combo.grid(row=1, column=1, sticky=W, padx=10, pady=8, ipady=4)

        Button(input_frame, text="Browse Excel File", font=("Segoe UI", 10, "bold"),
               bg="#1F4E79", fg="white", command=self.browse_file, relief=FLAT, cursor="hand2").grid(row=1, column=2, padx=10, pady=8)

        self.file_label = Label(input_frame, text="No file selected", bg="white", fg="#FF6B6B", font=("Segoe UI", 9, "italic"))
        self.file_label.grid(row=1, column=3, sticky=W, padx=5)

        Button(input_frame, text="Run Reconciliation", font=("Segoe UI", 11, "bold"),
               bg="#28A745", fg="white", command=self.run_threaded, relief=FLAT, cursor="hand2").grid(row=2, column=1, pady=12, padx=10, sticky=EW)

        self.progress = ttk.Progressbar(input_frame, orient=HORIZONTAL, length=400, mode="determinate")
        self.progress.grid(row=2, column=3, padx=10, pady=12, sticky=EW)

        # Correction workflow buttons (disabled until reconciliation completes)
        self.btn_download_draft = Button(input_frame, text="Download Draft Output", state=DISABLED,
                         command=self.download_draft_output)
        self.btn_download_draft.grid(row=3, column=0, pady=8)

        self.btn_download_correction_report = Button(input_frame, text="Download Correction Report", state=DISABLED,
                                command=self.download_correction_report)
        self.btn_download_correction_report.grid(row=3, column=1, pady=8)

        # note: upload correction and revised output features removed

        self.tree = ttk.Treeview(main_frame, show="headings")
        self.tree.pack(fill=BOTH, expand=True, pady=10)

        self.log_box = Text(main_frame, height=5, bg="#F9F9F9", font=("Consolas", 10))
        self.log_box.pack(fill=X)

        self.log("Dashboard Ready.")

    def browse_file(self):
        f = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if f:
            self.selected_file = f
            self.file_label.config(text=os.path.basename(f), fg="green")
            self.log(f"Selected file: {f}")

    def log(self, msg):
        ts = datetime.datetime.now().strftime("%H:%M:%S")
        self.log_box.insert(END, f"[{ts}] {msg}\n")
        self.log_box.see(END)

    def update_progress(self, val):
        self.progress["value"] = val
        self.root.update_idletasks()

    def run_threaded(self):
        if not self.selected_file:
            messagebox.showerror("Error", "Please select an Excel file.")
            return

        taxpayer_gstin = self.taxpayer_entry.get().strip()
        period = self.period_entry.get().strip()

        if taxpayer_gstin == "" or period == "":
            messagebox.showerror("Error", "Taxpayer GSTIN and Period are required.")
            return

        tolerance = self.tolerance_var.get()

        self.progress["value"] = 0
        self.log("Starting reconciliation...")

        thread = threading.Thread(target=self.run_reco, args=(taxpayer_gstin, period, tolerance))
        thread.daemon = True
        thread.start()

    def run_reco(self, taxpayer_gstin, period, tolerance):
        try:
            output_file, df = reconcile(
                self.selected_file,
                taxpayer_gstin,
                period,
                tolerance,
                self.current_user,
                progress_callback=lambda v: self.root.after(0, self.update_progress, v),
                log_callback=lambda m: self.root.after(0, self.log, m)
            )

            # This is the initial draft output (do not treat as final)
            self.draft_output_path = output_file
            self.output_file_path = None
            self.result_df = df

            # Enable correction workflow buttons (draft available)
            self.root.after(0, self._enable_correction_buttons)

            self.root.after(0, lambda: messagebox.showinfo(
                "Success",
                f"Reconciliation completed.\n\nOutput saved:\n{output_file}"
            ))

        except Exception as e:
            err_msg = str(e)
            self.root.after(0, lambda msg=err_msg: messagebox.showerror("Error", msg))

    def _enable_correction_buttons(self):
        # Enable draft + correction report + upload after draft available
        if self.draft_output_path and os.path.isfile(self.draft_output_path):
            try:
                self.btn_download_draft.config(state=NORMAL)
                self.btn_download_correction_report.config(state=NORMAL)
            except Exception:
                pass

        # Enable final output download only after final output exists
        if self.output_file_path and os.path.isfile(self.output_file_path):
            try:
                self.btn_download_final.config(state=NORMAL)
            except Exception:
                pass
        # note: corrected workflow removed, button logic gone

    def download_final_output(self):
        if not self.output_file_path or not os.path.isfile(self.output_file_path):
            messagebox.showerror("Error", "Final output not available.")
            return
        # Open containing folder for final output
        try:
            folder = os.path.dirname(self.output_file_path)
            os.startfile(folder)
        except Exception:
            messagebox.showinfo("Output Path", f"Final output file: {self.output_file_path}")

    def download_correction_report(self):
        if self.result_df is None:
            messagebox.showerror("Error", "Reco result not available to generate report.")
            return

        out_dir = os.path.dirname(self.selected_file) if self.selected_file else os.getcwd()

        def _gen():
            try:
                path, dfs = generate_correction_report(self.result_df, out_dir, log=self.log)
                self.root.after(0, lambda: messagebox.showinfo("Saved", f"Correction report saved:\n{path}"))
            except Exception as ex:
                self.root.after(0, lambda: messagebox.showerror("Error", str(ex)))

        threading.Thread(target=_gen, daemon=True).start()

    def download_draft_output(self):
        if not self.draft_output_path or not os.path.isfile(self.draft_output_path):
            messagebox.showerror("Error", "Draft output not available.")
            return
        try:
            folder = os.path.dirname(self.draft_output_path)
            os.startfile(folder)
        except Exception:
            messagebox.showinfo("Draft Path", f"Draft output file: {self.draft_output_path}")

    # correction uploading and revised logic removed

    # correction upload and processing removed

    # revised reconciliation logic removed

    def _show_correction_summary(self, mods, unmatched):
        """Display a Toplevel dialog listing applied modifications and unmatched rows, with option to save report."""
        try:
            win = Toplevel(self.root)
            win.title("Correction Summary")
            win.geometry("800x500")

            lbl = Label(win, text=f"Applied Modifications: {len(mods)}    Unmatched Rows: {len(unmatched)}", font=("Segoe UI", 10, "bold"))
            lbl.pack(pady=8)

            frame = Frame(win)
            frame.pack(fill=BOTH, expand=True, padx=8, pady=4)

            left = Frame(frame)
            left.pack(side=LEFT, fill=BOTH, expand=True)
            Label(left, text="Modifications (key, column, old -> new)").pack()
            txt_mod = Text(left, wrap=NONE)
            txt_mod.pack(fill=BOTH, expand=True)

            right = Frame(frame)
            right.pack(side=RIGHT, fill=BOTH, expand=True)
            Label(right, text="Unmatched Rows (sheet, key)").pack()
            txt_un = Text(right, wrap=NONE)
            txt_un.pack(fill=BOTH, expand=True)

            # Populate texts
            for m in mods:
                try:
                    txt_mod.insert(END, f"{m.get('key')} | {m.get('column')} : {m.get('old')} -> {m.get('new')}\n")
                except Exception:
                    continue

            for u in unmatched:
                if isinstance(u, dict):
                    txt_un.insert(END, f"{u.get('sheet')} | {u.get('key')}\n")
                else:
                    txt_un.insert(END, f"{str(u)}\n")

            btn_frame = Frame(win)
            btn_frame.pack(pady=8)

            def _save_report():
                try:
                    out_dir = os.path.dirname(self.selected_file) if self.selected_file else os.getcwd()
                    rpt_path = os.path.join(out_dir, f"Correction_Summary_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
                    df_mod = pd.DataFrame(mods)
                    df_un = pd.DataFrame(unmatched)
                    with pd.ExcelWriter(rpt_path, engine='openpyxl') as w:
                        df_mod.to_excel(w, sheet_name='Modifications', index=False)
                        df_un.to_excel(w, sheet_name='Unmatched', index=False)
                    messagebox.showinfo("Saved", f"Summary saved: {rpt_path}")
                except Exception as e:
                    messagebox.showerror("Save Failed", str(e))

            Button(btn_frame, text="Save Report", command=_save_report).pack(side=LEFT, padx=6)
            Button(btn_frame, text="Close", command=win.destroy).pack(side=RIGHT, padx=6)

        except Exception as ex:
            messagebox.showerror("Error", f"Failed to show summary: {ex}")


# ==========================================================
# MAIN (ROBUST TKINTER START)
# ==========================================================
if __name__ == "__main__":
    try:
        init_db()
        create_default_change_heading_file()

        root = Tk()
        root.withdraw()
        root.update()
        root.deiconify()

        app = GSTRecoApp(root)
        root.mainloop()

    except Exception as ex:
        print("APPLICATION CRASHED!")
        print(traceback.format_exc())
        messagebox.showerror("Fatal Error", str(ex))
